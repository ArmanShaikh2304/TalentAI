from __future__ import annotations
"""
Export API Endpoints — Export analysis results as CSV or professional Excel.
"""
import io
import csv
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.api.dependencies import get_current_user

router = APIRouter(prefix="/api/export", tags=["Export"])

# Import analysis store access
from app.api.analysis import _analysis_store


@router.get("/csv/{analysis_id}")
async def export_csv(analysis_id: str, _user: dict = Depends(get_current_user)):
    """Export analysis results as CSV."""
    if analysis_id not in _analysis_store:
        raise HTTPException(status_code=404, detail=f"Analysis '{analysis_id}' not found")

    analysis = _analysis_store[analysis_id]
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        "Rank", "Name", "Title", "Score (%)",
        "Semantic Similarity", "Technical Skills", "Experience Level",
        "Career Trajectory", "Domain Knowledge", "Behavioral Signals",
        "Red Flags", "Executive Summary", "Risk Assessment",
        "Long-term Potential", "Skills", "Experience (Years)",
    ])

    # Rows
    for candidate in analysis.ranked_candidates:
        dim_scores = {d.name: d.score for d in candidate.dimensions}
        writer.writerow([
            candidate.rank,
            candidate.name,
            candidate.title,
            round(candidate.total_score, 1),
            dim_scores.get("Semantic Similarity", 0),
            dim_scores.get("Technical Skills", 0),
            dim_scores.get("Experience Level", 0),
            dim_scores.get("Career Trajectory", 0),
            dim_scores.get("Domain Knowledge", 0),
            dim_scores.get("Behavioral Signals", 0),
            dim_scores.get("Red Flags", 0),
            candidate.explanation.executive_summary,
            candidate.explanation.risk_assessment.value,
            candidate.explanation.long_term_potential.value,
            ", ".join(candidate.skills),
            candidate.experience_years,
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=candidate_ranking_{analysis_id}.csv"},
    )


@router.post("/shortlisted/{analysis_id}")
async def export_shortlisted_excel(analysis_id: str, request: dict = None, _user: dict = Depends(get_current_user)):
    """
    Export only shortlisted candidates as a professional Excel (.xlsx) file.

    Accepts a JSON body with { "candidate_ids": ["id1", "id2", ...] } to export
    specific user-selected shortlisted candidates. If no candidate_ids provided,
    falls back to score-based threshold filtering.
    Returns 404 with message if no candidates are shortlisted.
    """
    if analysis_id not in _analysis_store:
        raise HTTPException(status_code=404, detail=f"Analysis '{analysis_id}' not found")

    analysis = _analysis_store[analysis_id]

    # Determine shortlisted candidates: prefer explicit IDs, fallback to score threshold
    candidate_ids = (request or {}).get("candidate_ids", []) if request else []

    if candidate_ids:
        # User has manually shortlisted specific candidates
        shortlisted = [
            c for c in analysis.ranked_candidates
            if c.candidate_id in candidate_ids
        ]
    else:
        # Fallback: use score threshold
        threshold = analysis.summary.shortlist_threshold
        shortlisted = [
            c for c in analysis.ranked_candidates
            if c.total_score >= threshold
        ]

    if not shortlisted:
        raise HTTPException(
            status_code=404,
            detail="No shortlisted candidates available.",
        )

    # ─── Build Professional Excel Workbook ───
    wb = Workbook()
    ws = wb.active
    ws.title = "Shortlisted Candidates"

    # ─── Define Styles ───
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_border = Border(
        bottom=Side(style="thin", color="334155"),
        right=Side(style="thin", color="334155"),
    )

    # Alternating row fills
    row_fill_even = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    cell_font = Font(name="Calibri", size=10)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    cell_border = Border(
        bottom=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
    )

    # Score highlight (green for high scores)
    score_font_high = Font(name="Calibri", bold=True, size=10, color="059669")
    score_font_medium = Font(name="Calibri", bold=True, size=10, color="D97706")
    score_font_low = Font(name="Calibri", size=10, color="DC2626")

    # ─── Column Headers ───
    headers = [
        "Candidate ID",
        "Rank",
        "Full Name",
        "Title",
        "Email",
        "Phone Number",
        "Experience (Years)",
        "Skills",
        "Education",
        "Certifications",
        "Total Score (%)",
        "JD Match Score",
        "AI Rank",
        "Shortlisted Status",
        "Recommendation",
        "Risk Assessment",
        "Long-term Potential",
        "Executive Summary",
        "Strengths",
        "Considerations",
        "Upload Date",
    ]

    # Dimension columns (dynamic based on candidate data)
    dim_names = []
    if shortlisted:
        dim_names = [d.name for d in shortlisted[0].dimensions]
        headers.extend([f"{name} (%)" for name in dim_names])

    # Write header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    # ─── Write Data Rows ───
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    for row_idx, candidate in enumerate(shortlisted, 2):
        dim_scores = {d.name: d.score for d in candidate.dimensions}

        # Semantic similarity / JD Match score
        jd_match = dim_scores.get("Semantic Similarity", candidate.total_score)

        # Compile strengths and considerations
        strengths_text = "; ".join([s.text for s in candidate.explanation.strengths[:3]]) if candidate.explanation.strengths else "—"
        considerations_text = "; ".join([d.text for d in candidate.explanation.development_areas[:2]]) if candidate.explanation.development_areas else "—"

        # Recommendation text
        if candidate.total_score >= 90:
            recommendation = "Highly Recommended — Strong Fit"
        elif candidate.total_score >= 80:
            recommendation = "Recommended — Good Fit"
        else:
            recommendation = "Consider — Moderate Fit"

        row_data = [
            candidate.candidate_id,
            candidate.rank,
            candidate.name,
            candidate.title,
            "",  # Email — not in current data model, placeholder
            "",  # Phone — not in current data model, placeholder
            candidate.experience_years,
            ", ".join(candidate.skills),
            candidate.education or "—",
            ", ".join(candidate.certifications) if candidate.certifications else "—",
            round(candidate.total_score, 1),
            round(jd_match, 1),
            candidate.rank,
            "✅ Shortlisted",
            recommendation,
            candidate.explanation.risk_assessment.value.title(),
            candidate.explanation.long_term_potential.value.title(),
            candidate.explanation.executive_summary,
            strengths_text,
            considerations_text,
            today_str,
        ]

        # Add dimension scores
        for dim_name in dim_names:
            row_data.append(round(dim_scores.get(dim_name, 0), 1))

        # Write cells
        row_fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = cell_border
            cell.fill = row_fill

            # Special formatting for score columns
            if col_idx == 11:  # Total Score
                score = value if isinstance(value, (int, float)) else 0
                if score >= 90:
                    cell.font = score_font_high
                elif score >= 80:
                    cell.font = score_font_medium
                else:
                    cell.font = score_font_low

    # ─── Auto-adjust Column Widths ───
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = len(str(headers[col_idx - 1]))

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), 50))

        ws.column_dimensions[column_letter].width = max_length + 4

    # ─── Freeze First Row & Add Auto-Filter ───
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ─── Save to BytesIO ───
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename with date
    date_str = datetime.now(timezone.utc).strftime("%Y_%m_%d")
    filename = f"Shortlisted_Candidates_{date_str}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
